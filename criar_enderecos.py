import pyautogui
import time
import openpyxl
import pyperclip
from pathlib import Path
import tkinter as tk
import threading
from tkinter import messagebox
import ttkbootstrap as ttk


# Variáveis globais
script_rodando = False

# Consumir planilha
workbook = openpyxl.load_workbook('planilha_para_enderecos.xlsx', data_only=True)
sheet_enderecamento = workbook['Endereçamento']

# Diretório onde o script está localizado
#diretorio_atual = Path(__file__).parent  # Caminho base do script

# Caminho para as imagens
img_caminho_inicio = r'C:\Users\img/inicio.png'
img_caminho_cadastro = r'C:\Users\img/cadastros.png'
img_caminho_logistica = r'C:\Users\img/logisticas.png'
img_caminho_enderecamentos = r'C:\Users\img/enderecamentos.png'
img_caminho_salvar = r'C:\Users\img/salvar.png'
img_caminho_novo = r'C:\Users\img/novo.png'



# Verificando se as imagens existem antes de usá-las
#for img in [img_caminho_inicio, img_caminho_cadastro, img_caminho_logistica, img_caminho_enderecamentos]:
 #   if not img.exists():
  #      raise FileNotFoundError(f"A imagem {img} não foi encontrada!")

def iniciar_script():
    global script_rodando
    script_rodando = True

    def executar_script():
        try:
        # Loop para Iniciar o Script
            procurar = True
            while procurar:
                try:
                    if script_rodando is False:
                        log("Script interrompido.")
                        break
                    pyautogui.locateCenterOnScreen(str(img_caminho_inicio), confidence=0.7)
                    log(f'Encontrei a janela, entrando...!')
                    procurar = False
                except:
                    if script_rodando is False:
                        log("Script interrompido.")
                        break
                    time.sleep(1)
                    log('Imagem não encontrada, tentando novamente...')
                    
        # Função para entrar na janela de endereços
            def entrar_na_janela_enderecos():
                # Clicar no Meno Cadastros
                cadastros = pyautogui.locateCenterOnScreen(str(img_caminho_cadastro), confidence=0.7)
                time.sleep(0.3)
                pyautogui.click(cadastros.x, cadastros.y)
                time.sleep(0.3)

                # Clicar no Sub-Menu Logistica
                logisticas = pyautogui.locateCenterOnScreen(str(img_caminho_logistica), confidence=0.7)
                time.sleep(0.3)
                pyautogui.click(logisticas.x, logisticas.y)

                # Clicar no Sub-Menu Endereçamento
                enderecamentos = pyautogui.locateCenterOnScreen(str(img_caminho_enderecamentos), confidence=0.7)
                time.sleep(0.3)
                pyautogui.click(enderecamentos.x, enderecamentos.y)

            entrar_na_janela_enderecos()
            log('Econtrei a janela com sucesso! \n')


            def digitar_campos_enderecos():
                log("o Script sera iniciado.\n")

                # Encontra a última linha com dados na coluna A
                ultima_linha_com_dados = None
                log('Processando total de linhas...')
                # Itera de baixo para cima na coluna A
                for linha in reversed(list(sheet_enderecamento.iter_rows(min_col=1, max_col=1))):
                    if linha[0].value is not None and str(linha[0].value).strip() != "":  # Verifica se não está vazio
                        ultima_linha_com_dados = linha[0].row  # Armazena a linha onde encontrou o valor
                        break  # Para a iteração assim que encontrar a última linha com dados

                log(f'{ultima_linha_com_dados} linhas econtradas\n')

                for i, linha in enumerate(sheet_enderecamento.iter_rows(min_row=2, min_col=2), start=2):
                    # Verifica se o script foi interrompido
                    if linha == ultima_linha_com_dados:  
                        messagebox.showinfo("Sucesso", "O script foi executado com sucesso.")
                        break
                    elif script_rodando is False:
                        log("Script interrompido.")
                        messagebox.showerror("Error", "O script foi interrompido.")
                        break
                    try:
                        # Colocando a Area
                        log(f'Processando linha {i} de {ultima_linha_com_dados} linhas')
                        time.sleep(0.5)
                        pyautogui.hotkey('tab')
                        pyautogui.hotkey('enter')
                        time.sleep(0.1)
                        area_planilha = sheet_enderecamento['A2'].value
                        pyperclip.copy(area_planilha)
                        pyautogui.hotkey('ctrl', 'v')
                        time.sleep(0.1)
                        pyautogui.hotkey('enter')     
                        time.sleep(0.1)
                        pyautogui.hotkey('enter')
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Fileira
                        fileira = linha[0].value
                        fileira = str(fileira)
                        pyautogui.typewrite(fileira)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Coluna
                        coluna = linha[1].value
                        coluna = str(coluna)
                        pyautogui.typewrite(coluna)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Andar
                        andar = linha[2].value
                        andar = str(andar)
                        pyautogui.typewrite(andar)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Posição
                        posicao = linha[3].value
                        posicao = str(posicao)
                        pyautogui.typewrite(posicao)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Tamanho
                        tamanho = linha[4].value
                        pyautogui.typewrite(tamanho)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Compartilhado
                        compartilhado = linha[5].value
                        pyautogui.typewrite(compartilhado)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Máximo Produto
                        max_produtos = linha[6].value
                        max_produtos = str(max_produtos)
                        pyautogui.typewrite(max_produtos)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Class. End.
                        pyautogui.typewrite('B')
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Class. Validade
                        pyautogui.typewrite('A')
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Temporario
                        pyperclip.copy("Não")
                        pyautogui.hotkey("ctrl", "v")  
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Protegido
                        pyperclip.copy("Não")
                        pyautogui.hotkey("ctrl", "v")   
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Temperatura  
                        pyperclip.copy("Não")
                        pyautogui.hotkey("ctrl", "v")
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Especial
                        pyperclip.copy("Não")
                        pyautogui.hotkey("ctrl", "v")
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
                        # Ordenação
                        pyautogui.hotkey('tab')
                        pyautogui.hotkey('tab')
                        pyautogui.hotkey('tab')
                        pyautogui.hotkey('tab')
                        pyautogui.typewrite('0')
                        time.sleep(0.3)

                        # Salvar endereço atual
                        salvar = pyautogui.locateCenterOnScreen(str(img_caminho_salvar), confidence=0.5)
                        pyautogui.click(salvar.x, salvar.y)
                        pyautogui.click(salvar.x, salvar.y)
                        time.sleep(0.7)
                        pyautogui.hotkey('enter')
                        time.sleep(0.2)

                        # Iniciar nova linha
                        time.sleep(0.5)
                        novo = pyautogui.locateCenterOnScreen(str(img_caminho_novo), confidence=0.7)
                        pyautogui.click(novo.x, novo.y)
                        pyautogui.click(novo.x, novo.y)
                        time.sleep(0.5)

                    # Interrompe o loop se ocorrer um erro
                    except:
                        log(f'Erro ao processar linha {i}')
                        break   
            digitar_campos_enderecos()
        except:
            messagebox.showerror("Error", "O script foi interrompido.")


        

    # Executar o script em uma thread separada
    threading.Thread(target=executar_script).start()

def parar_script():
    global script_rodando
    script_rodando = False

# Função para logar as mensagens no widget Text
def log(msg):
    text_log.config(state=tk.NORMAL)  # Habilitar edição para inserir o texto
    text_log.insert(tk.END, f"{msg}\n")
    text_log.yview(tk.END)  # Descer o scroll automaticamente  
    text_log.config(state=tk.DISABLED)  # Desabilitar edição após a inserção 

# Criar janela
janela = ttk.Window(
    themename = "darkly"
    ) 
janela.title("Criar Endereços")
janela.geometry("600x400")
janela.resizable(False, False)


# Títulos e informações
label_titulo = tk.Label(
    janela,
    text = "Criar Endereços",
    font = ("Helvetica", 16, "bold"),
    fg = "white",
    bg = "#2c3e50",
)

label_titulo.pack(pady=10)

# Botões
btn_iniciar = ttk.Button(
    janela, text = "Iniciar Script", 
    bootstyle = "success", 
    command = iniciar_script
    )

btn_iniciar.pack(
    pady = 10
    )

btn_parar = ttk.Button(
    janela, text="Parar Script", 
    bootstyle = "danger", 
    command = parar_script
    )

btn_parar.pack(
    pady = 10
    )

# Área de texto para log
text_log = tk.Text(
    janela,
    height = 20, 
    width = 90
    )

text_log.pack(
    pady = 20
    )
# Desabilitar edição
text_log.config(
    state = tk.DISABLED
    ) 

# Função para permitir a escrita no Text
def enable_text():
    text_log.config(state=tk.NORMAL)

# Rodar a interface
janela.mainloop()
