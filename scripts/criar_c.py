import time
import os
import pyautogui
import openpyxl
from tkinter import messagebox
from interface.log import log


# Variáveis globais
script_rodando = False

# Consumir planilha
workbook = openpyxl.load_workbook('planilha_para_enderecos.xlsx', data_only=True)
aba_panilha_enderecamento = workbook['Endereçamento']


# Caminho para as imagens
img_caminho_inicio = r'C:\Users\img/inicio.png'
img_caminho_cadastro = r'C:\Users\img/cadastros.png'
img_caminho_logistica = r'C:\Users\img/logisticas.png'
img_caminho_enderecamentos = r'C:\Users\img/enderecamentos.png'
img_caminho_salvar = r'C:\Users\img/salvar.png'
img_caminho_novo = r'C:\Users\img/novo.png'
img_caminho_confirmacao = r'C:\Users\img/confirmacao.png'



# Validação inicial das imagens
def verificar_imagens():
    imagens = [
        img_caminho_inicio, img_caminho_cadastro, img_caminho_logistica,
        img_caminho_enderecamentos, img_caminho_salvar, img_caminho_novo,
        img_caminho_confirmacao
    ]
    for img in imagens:
        if not os.path.exists(img):
            raise FileNotFoundError(f"A imagem {img} não foi encontrada!")

# Validando as imagens antes de iniciar
try:
    verificar_imagens()
except FileNotFoundError as e:
    messagebox.showerror("Erro", str(e))
    exit(1)

# Iniciar Script 
def iniciar_script():
    global script_rodando
    script_rodando = True

    def executar_script():
        try:
        # Loop para Iniciar o Script
            procurar = True
            while procurar:
                try:
                    if script_rodando is False:
                        log("Script interrompido.")
                        break
                    pyautogui.locateCenterOnScreen(str(img_caminho_inicio), confidence=0.7)
                    log(f'Encontrei a janela, entrando...!')
                    procurar = False
                except:
                    if script_rodando is False:
                        log("Script interrompido.")
                        break
                    time.sleep(1)
                    log('Imagem não encontrada, tentando novamente...')
                    
            # Clicar no Meno Cadastros
            cadastros = pyautogui.locateCenterOnScreen(str(img_caminho_cadastro), confidence=0.7)
            logisticas = pyautogui.locateCenterOnScreen(str(img_caminho_logistica), confidence=0.7)
            enderecamentos = pyautogui.locateCenterOnScreen(str(img_caminho_enderecamentos), confidence=0.7)

            menus = [cadastros, logisticas, enderecamentos]

            for menu in menus:
                time.sleep(0.3)
                pyautogui.click(menu.x, menu.y)
                time.sleep(0.3)
                


            log('Econtrei a janela com sucesso! \n')
            log("o Script sera iniciado.\n")

            # Encontra a última linha com dados na coluna A
            ultima_linha_com_dados = None

            log('Processando total de linhas...')

            # Itera de baixo para cima na coluna A
            for linha in reversed(list(aba_panilha_enderecamento.iter_rows(min_col=1, max_col=1))):
                if linha[0].value is not None and str(linha[0].value).strip() != "":
                    ultima_linha_com_dados = linha[0].row
                    break  

            log(f'{ultima_linha_com_dados} linhas econtradas\n')

            for i, linha in enumerate(aba_panilha_enderecamento.iter_rows(min_row=2, min_col=2), start=2):   
                try:
                    # Colocando a Area e validando
                    time.sleep(0.5)
                    pyautogui.hotkey('tab')
                    pyautogui.hotkey('enter')
                    time.sleep(0.1)

                    area = str(aba_panilha_enderecamento['A2'].value).strip()
                    if len(area) == 1 and area.isdigit():
                        area = f"0{area}"
                    else:
                        log(f"Erro na linha {i}: 'area' inválida ({area}). Pulando linha...")
                        continue


                    pyautogui.typewrite(area.lstrip('0')) 
                    pyautogui.typewrite(area)
                    time.sleep(0.1)
                    pyautogui.hotkey('enter')     
                    time.sleep(0.1)
                    pyautogui.hotkey('enter')
                    pyautogui.hotkey('tab')
                    time.sleep(0.1)

                    # Posições e Validações
                    fileira = linha[0].value
                    if not validar_fileira(fileira):
                        log(f"Erro na linha {i}: 'fileira' inválida ({fileira}). Pulando linha...")
                        continue

                    coluna = linha[1].value
                    if not validar_coluna(coluna):
                        log(f"Erro na linha {i}: 'coluna' inválida ({coluna}). Pulando linha...")
                        continue

                    andar = linha[2].value
                    if not validar_andar(andar):
                        log(f"Erro na linha {i}: 'andar' inválido ({andar}). Pulando linha...")
                        continue
                    
                    posicao = linha[3].value
                    if not validar_posicao(posicao):
                        log(f"Erro na linha {i}: 'posição' inválida ({posicao}). Pulando linha...")
                        continue

                    tamanho = linha[4].value
                    compartilhado = linha[5].value
                    max_produtos = linha[6].value

                    
                    # Processamento da linha (caso todas as validações sejam bem-sucedidas)
                    log(f"Processando linha {i}: fileira={fileira}, coluna={coluna}, posição={posicao} de {ultima_linha_com_dados}")


                    campos = [fileira, coluna, andar, posicao, tamanho, compartilhado, max_produtos, 'B', 'A', 'Não', 'Não', 'Não', 'Não']

                    for campo in campos:
                        campo = str(campo)
                        pyautogui.typewrite(campo)
                        pyautogui.hotkey('tab')
                        time.sleep(0.1)
    
                    # Ordenação
                    for _ in range(4): 
                        pyautogui.hotkey('tab')
                        pyautogui.typewrite('0')
                        time.sleep(0.3)


                    # Salvar endereço atual
                    salvar = pyautogui.locateCenterOnScreen(str(img_caminho_salvar), confidence=0.5)
                    pyautogui.doubleClick(salvar.x, salvar.y)
                    time.sleep(0.7)
                    pyautogui.hotkey('enter')
                    time.sleep(0.2)

                   # Confirmação de operação
                    procurar_confirmacao = True
                    tempo_inicial = time.time() 
                    timeout = 10  

                    while procurar_confirmacao:
                        if time.time() - tempo_inicial > timeout:
                            log(f"Confirmação para a linha {i} não encontrada dentro do tempo limite.")
                            break  

                        try:
                            confirmacao = pyautogui.locateCenterOnScreen(str(img_caminho_confirmacao), confidence=0.7)
                            if confirmacao:
                                log(f"Linha {i} salva com sucesso!")
                                procurar_confirmacao = False 
                        except Exception as e:
                            log(f"Erro ao procurar confirmação: {e}") 
                        time.sleep(1)  
                    

                    if i == ultima_linha_com_dados:
                        log("Todas as linhas foram processadas com sucesso.")
                        messagebox.showinfo("Sucesso", "O script foi executado com sucesso.")
                        break


                    # Iniciar nova linha
                    time.sleep(0.5)
                    novo = pyautogui.locateCenterOnScreen(str(img_caminho_novo), confidence=0.7)
                    pyautogui.doubleClick(novo.x, novo.y)
                    time.sleep(0.5)
                    
                    # Interrompe o loop se ocorrer um erro
                except Exception as e:
                    log(f"Erro ao processar linha {i}: {e}")
                    messagebox.showerror("Erro", f"Erro na linha {i}: {e}")
                    break  
                
        except:
            messagebox.showerror("Error", "O script foi interrompido.")
    executar_script()

# Funções
# Validar campos

def validar_fileira(fileira):
    if isinstance(fileira, int):
        fileira = f"{fileira:02}"
    elif isinstance(fileira, str) and len(fileira) == 2 and fileira.isdigit():
        pass  
    else:
        return False
    return True

def validar_coluna(coluna):
    if isinstance(coluna, int):
        coluna = f"{coluna:02}" 
    elif isinstance(coluna, str) and len(coluna) == 2 and coluna.isdigit():
        pass  #
    else:
        return False 
    return True

def validar_andar(andar):
    andar = str(andar).strip()
    if andar.isdigit() and len(andar) == 1:
        pass 
    elif andar.isdigit() and len(andar) == 2 and andar[0] == '0' and andar[1] in '123456789':
        andar = andar[1]  
    else:
        return False
    return True

def validar_posicao(posicao):
    if isinstance(posicao, int): 
        posicao = f"{posicao:02}" 
    elif isinstance(posicao, str) and len(posicao) == 2 and posicao.isdigit():
        pass 
    else:
        return False 
    return True


def parar_script(text_log):
    global script_rodando
    if script_rodando:
        script_rodando = False
        log("Script parado.", text_log)
    else:
        log("Nenhum script em execução.", text_log)