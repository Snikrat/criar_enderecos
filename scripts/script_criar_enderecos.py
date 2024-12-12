#-*- coding: utf-8 -*-
import os
import json
import openpyxl
import time
import cv2
import pyautogui
import pyperclip
from tkinter import messagebox
from interface.log import *



def iniciar_script_criar_enderecos(text_log):
    # Função para carregar o caminho da planilha
    def carregar_caminho_planilha():
        try:
            with open("configuracoes_imagens.json", "r", encoding="utf-8") as file:
                configuracoes = json.load(file)
                caminho_planilha = configuracoes.get("planilhas", {}).get("caminho_planilha", "")
                
                if not caminho_planilha:
                    messagebox.showerror("Erro", "Caminho da planilha não encontrado no arquivo JSON.")
                    return None
                
                if not os.path.exists(caminho_planilha):
                    messagebox.showerror("Erro", f"A planilha não foi encontrada: {caminho_planilha}")
                    return None
                
                return caminho_planilha
        
        except json.JSONDecodeError:
            messagebox.showerror("Erro", "Erro ao decodificar o JSON. Verifique o formato do arquivo.")
            return None
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado ao carregar o caminho da planilha: {str(e)}")
            return None

    # Função para carregar os caminhos das imagens
    def carregar_caminhos_imagens():
        try:
            with open("configuracoes_imagens.json", "r", encoding="utf-8") as file:
                configuracoes = json.load(file)
                imagens = configuracoes.get("imagens", {})
                
                # Dicionário para armazenar os caminhos
                caminhos_imagens = {
                    "img_caminho_inicio": imagens.get("img_caminho_inicio", ""),
                    "img_caminho_cadastro": imagens.get("img_caminho_cadastro", ""),
                    "img_caminho_logistica": imagens.get("img_caminho_logistica", ""),
                    "img_caminho_enderecamentos": imagens.get("img_caminho_enderecamentos", ""),
                    "img_caminho_salvar": imagens.get("img_caminho_salvar", ""),
                    "img_caminho_novo": imagens.get("img_caminho_novo", ""),
                    "img_caminho_confirmacao": imagens.get("img_caminho_confirmacao", "")
                }

                # Validar a existência de cada caminho
                for chave, caminho in caminhos_imagens.items():
                    if not os.path.exists(caminho):
                        messagebox.showerror("Erro", f"Imagem não encontrada para {chave}: {caminho}")

                return caminhos_imagens
        
        except json.JSONDecodeError:
            messagebox.showerror("Erro", "Erro ao decodificar o JSON. Verifique o formato do arquivo.")
            return {}
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado ao carregar caminhos das imagens: {str(e)}")
            return {}

    # Uso das funções
    caminho_planilha = carregar_caminho_planilha()
    if caminho_planilha:
        try:
            # Abrir a planilha
            planilha_enderecos = openpyxl.load_workbook(caminho_planilha, data_only=True)
            aba_planilha_enderecamento = planilha_enderecos['Endereçamento']
            log("Planilha carregada com sucesso.", text_log)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar a planilha: {str(e)}")

    caminhos_imagens = carregar_caminhos_imagens()

    img_caminho_inicio = caminhos_imagens.get("img_caminho_inicio", "") 
    img_caminho_cadastro = caminhos_imagens.get("img_caminho_cadastro", "") 
    img_caminho_logistica= caminhos_imagens.get("img_caminho_logistica", "") 
    img_caminho_enderecamentos= caminhos_imagens.get("img_caminho_enderecamentos", "") 
    img_caminho_salvar= caminhos_imagens.get("img_caminho_salvar", "") 
    img_caminho_novo= caminhos_imagens.get("img_caminho_novo", "") 
    img_caminho_confirmacao= caminhos_imagens.get("img_caminho_confirmacao", "") 


    # Loop para Iniciar o Script
    def executando_loop():
            esperar_wms_carregar = True
            while esperar_wms_carregar:
                try:
                    pyautogui.locateCenterOnScreen((img_caminho_inicio), confidence=0.7)
                    log('O WMS carregou, inicianddo...', text_log)
                    esperar_wms_carregar = False
                except:
                    time.sleep(1)
                    log('Esperando o WMS carregar...', text_log)
    executando_loop()

    # Entrar na janela de endereços
    # Clicando em cadastros
    cadastros = pyautogui.locateCenterOnScreen(str(img_caminho_cadastro), confidence=0.7, grayscale=True)
    time.sleep(0.3)
    pyautogui.click(cadastros.x, cadastros.y)
    # Clicando em logisticas
    logisticas = pyautogui.locateCenterOnScreen(str(img_caminho_logistica), confidence=0.7, grayscale=True)
    time.sleep(0.3)
    pyautogui.click(logisticas.x, logisticas.y)
    # Clicando em enderecamentos
    enderecamentos = pyautogui.locateCenterOnScreen(str(img_caminho_enderecamentos), confidence=0.7, grayscale=True)
    time.sleep(0.3)
    pyautogui.click(enderecamentos.x, enderecamentos.y)

    log('Econtrei a janela com sucesso! \n', text_log)

    # Encontrando a última linha com dados na coluna A
    ultima_linha_com_dados = None
    log('Processando total de linhas...', text_log)

    for linha in reversed(list(aba_planilha_enderecamento.iter_rows(min_col=1, max_col=1))):
        if linha[0].value is not None and str(linha[0].value).strip() != "":  
            ultima_linha_com_dados = linha[0].row  
            break  
            
    log(f'{ultima_linha_com_dados} linhas econtradas\n', text_log)

    log("O Script sera iniciado.\n", text_log)
    
    for i, linha in enumerate(aba_planilha_enderecamento.iter_rows(min_row=2, min_col=2), start=2): 
        try: 
            # Colocando a Area e validando
            time.sleep(0.5)
            pyautogui.hotkey('tab')
            pyautogui.hotkey('enter')
            time.sleep(0.1)

            area = str(aba_planilha_enderecamento['A2'].value).strip()
        
            pyperclip.copy(area)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.1)
            pyautogui.hotkey('enter')     
            time.sleep(0.1)
            pyautogui.hotkey('enter')
            pyautogui.hotkey('tab')
            time.sleep(0.1)
            

        # Posições e Validações
            fileira = linha[0].value
            if not validar_fileira(fileira):
                log(f"Erro na linha {i}: 'fileira' inválida ({fileira}). Pulando linha...", text_log)
                continue

            coluna = linha[1].value
            if not validar_coluna(coluna):
                log(f"Erro na linha {i}: 'coluna' inválida ({coluna}). Pulando linha...", text_log)
                continue

            andar = linha[2].value
            if not validar_andar(andar):
                log(f"Erro na linha {i}: 'andar' inválido ({andar}). Pulando linha...", text_log)
                continue
                    
            posicao = linha[3].value
            if not validar_posicao(posicao):
                log(f"Erro na linha {i}: 'posição' inválida ({posicao}). Pulando linha...", text_log)
                continue

            tamanho = linha[4].value
            compartilhado = linha[5].value
            max_produtos = linha[6].value

             # Processamento da linha (caso todas as validações sejam bem-sucedidas)
            log(f"Processando linha {i}: fileira={fileira}, coluna={coluna}, posição={posicao} de {ultima_linha_com_dados}", text_log)

            campos = [fileira, coluna, andar, posicao, tamanho, compartilhado, max_produtos, 'B', 'A', 'Não', 'Não', 'Não', 'Não']

            for campo in campos:
                campo = str(campo)
                pyperclip.copy(campo)
                pyautogui.hotkey('ctrl', 'v')
                pyautogui.hotkey('tab')
                time.sleep(0.1)
    
            # Ordenação
            for _ in range(4): 
                pyautogui.hotkey('tab')
            pyautogui.typewrite('0')
            time.sleep(0.3)

            # Salvar endereço atual
            salvar = pyautogui.locateCenterOnScreen(str(img_caminho_salvar), confidence=0.5)
            pyautogui.doubleClick(salvar.x, salvar.y)
            time.sleep(0.7)
            pyautogui.hotkey('enter')
            time.sleep(0.2)


        except:
            print("Ollá")

# Validar campos
def validar_fileira(fileira):
    if isinstance(fileira, int):
        fileira = f"{fileira:02}"
    elif isinstance(fileira, str) and len(fileira) == 2 and fileira.isdigit():
        pass  
    else:
        return False
    return True

def validar_coluna(coluna):
    if isinstance(coluna, int):
        coluna = f"{coluna:02}" 
    elif isinstance(coluna, str) and len(coluna) == 2 and coluna.isdigit():
        pass  #
    else:
        return False 
    return True

def validar_andar(andar):
    andar = str(andar).strip()
    if andar.isdigit() and len(andar) == 1:
        pass 
    elif andar.isdigit() and len(andar) == 2 and andar[0] == '0' and andar[1] in '123456789':
        andar = andar[1]  
    else:
        return False
    return True

def validar_posicao(posicao):
    if isinstance(posicao, int): 
        posicao = f"{posicao:02}" 
    elif isinstance(posicao, str) and len(posicao) == 2 and posicao.isdigit():
        pass 
    else:
        return False 
    return True


# def parar_script(text_log):
#     global script_rodando
#     if script_rodando:
#         script_rodando = False
#         log("Script parado.", text_log)
#     else:
#         log("Nenhum script em execução.", text_log)    
